import os
import pandas as pd
from openpyxl import load_workbook
import xlrd
from docx import Document
import datetime  # 新增导入，用于获取当前时间

# 让用户输入要查询的关键字和目录地址
keywords = input("请输入要查询的关键字，以空格键隔开: ").split()
excel_dir = input("请输入要查找的目录地址: ")
# 以“results+当前时间”命名输出文件
now = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
output_file = f"results_{now}.xlsx"

# 初始化结果列表
results = []

# 处理 Excel 文件
def process_excel(file_path):
    try:
        if file_path.endswith('.xlsx'):
            # 处理 .xlsx 文件
            wb = load_workbook(file_path, read_only=True)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value and any(key in str(cell.value) for key in keywords):
                            result = {
                                "文件类型": "Excel",
                                "文件名": os.path.basename(file_path),
                                "文件路径": f'=HYPERLINK("{file_path}", "{file_path}")',
                                "工作表": sheet_name,
                                "位置": cell.coordinate,
                                "内容": cell.value
                            }
                            results.append(result)
            wb.close()
        elif file_path.endswith('.xls'):
            # 处理 .xls 文件
            wb = xlrd.open_workbook(file_path)
            for sheet_name in wb.sheet_names():
                sheet = wb.sheet_by_name(sheet_name)
                for row in range(sheet.nrows):
                    for col in range(sheet.ncols):
                        cell_value = sheet.cell_value(row, col)
                        if cell_value and any(key in str(cell_value) for key in keywords):
                            cell_coordinate = f"{chr(65 + col)}{row + 1}"
                            result = {
                                "文件类型": "Excel",
                                "文件名": os.path.basename(file_path),
                                "文件路径": f'=HYPERLINK("{file_path}", "{file_path}")',
                                "工作表": sheet_name,
                                "位置": cell_coordinate,
                                "内容": cell_value
                            }
                            results.append(result)
    except Exception as e:
        error_msg = f"处理文件 {os.path.basename(file_path)} 时出错: {str(e)}, 路径: {file_path}"
        print(error_msg)

# 处理 Word 文件
def process_word(file_path):
    try:
        doc = Document(file_path)
        for para in doc.paragraphs:
            if any(key in para.text for key in keywords):
                result = {
                    "文件类型": "Word",
                    "文件名": os.path.basename(file_path),
                    "文件路径": f'=HYPERLINK("{file_path}", "{file_path}")',
                    "工作表": "N/A",
                    "位置": "段落",
                    "内容": para.text
                }
                results.append(result)
    except Exception as e:
        error_msg = f"处理文件 {os.path.basename(file_path)} 时出错: {str(e)}, 路径: {file_path}"
        print(error_msg)

# 遍历目录及其子目录下所有 Word 和 Excel 文件
for root, dirs, files in os.walk(excel_dir):
    for file in files:
        # 过滤掉隐藏文件和临时文件
        if file.startswith(('.', '~$')):
            continue
        if file.endswith((".xlsx", ".xls")):
            file_path = os.path.join(root, file)
            process_excel(file_path)
        elif file.endswith(".docx"):
            file_path = os.path.join(root, file)
            process_word(file_path)

# 将结果保存到 Excel 文件
if results:
    df = pd.DataFrame(results)
    df.to_excel(output_file, index=False)
    print(f"搜索结果已保存到 {output_file}")
else:
    print("未找到匹配的结果。")